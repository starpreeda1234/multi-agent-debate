import sys, io, time, json, os
from datetime import datetime
from openai import OpenAI
import gradio as gr

# ── GLOBAL CONTROLLER FOR STREAM STOP ──
debate_running_flag = True

PROVIDERS = {
    "openai": {"base_url": "https://api.openai.com/v1", "models": ["gpt-4o", "gpt-4o-mini"]},
    "alibaba": {"base_url": "https://ws-tibabjyugiz6yoo5.ap-southeast-1.maas.aliyuncs.com/compatible-mode/v1", "models": ["qwen-plus", "qwen-turbo", "qwen-max"]},
    "anthropic": {"base_url": "https://api.anthropic.com/v1", "models": ["claude-3-5-sonnet-20241022"]},
    "groq": {"base_url": "https://api.groq.com/openai/v1", "models": ["llama3-70b-8192"]},
    "google_gemini": {"base_url": "https://generativelanguage.googleapis.com/v1beta/openai", "models": ["gemini-1.5-pro", "gemini-1.5-flash"]},
    "deepseek": {"base_url": "https://api.deepseek.com/v1", "models": ["deepseek-chat"]},
    "custom": {"base_url": "", "models": []}
}

DEFAULT_AGENTS = [
    {"name": "ดร.อนันต์ (วิศวกรชลศาสตร์)", "icon": "🤖", "provider": "alibaba", "api_key": "", "model": "qwen-plus", "temperature": 0.7, "top_p": 0.9, "role": "ผู้เชี่ยวชาญด้านวิศวกรรมชลศาสตร์ ประสบการณ์บริหารน้ำ 20 ปี เน้นโครงสร้างพื้นฐาน"},
    {"name": "ดร.ศิริพร (นักเศรษฐศาสตร์เมือง)", "icon": "💡", "provider": "alibaba", "api_key": "", "model": "qwen-plus", "temperature": 0.6, "top_p": 0.85, "role": "นักเศรษฐศาสตร์เมือง เชี่ยวชาญด้านต้นทุนความเสียหายและงบประมาณเยียวยา"},
    {"name": "ศ.กิตติ (นักสิ่งแวดล้อม)", "icon": "💵", "provider": "alibaba", "api_key": "", "model": "qwen-plus", "temperature": 0.8, "top_p": 0.95, "role": "ผู้เชี่ยวชาญด้านสิ่งแวดล้อมและ Climate Change"},
]

EMOJI_CHOICES = ["🤖", "🏛️", "🌾", "🏥", "📢", "👨‍💼", "👩‍💼", "👨‍🔬", "🌳", "🏭", "📊", "💡", "🎯", "🛡️", "🚀", "💵"]

CUSTOM_CSS = """
.gradio-container { font-family: 'Inter', sans-serif; max-width: 1350px !important; margin: 0 auto; }
.stat-card { padding: 14px; border-radius: 8px; text-align: center; border: 1px solid #e5e7eb; }
.control-box { padding: 20px; border-radius: 12px; margin-bottom: 15px; border: 1px solid #e5e7eb; }
"""

def _write_excel(records, topic, rounds, path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError: return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "บทสนทนา"
    metadata = [("หัวข้อ:", topic), ("วันเวลา:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")), ("รอบรวม:", rounds)]
    for r_idx, vals in enumerate(metadata, 1):
        ws.cell(row=r_idx, column=1, value=vals[0]).font = Font(bold=True)
        ws.cell(row=r_idx, column=2, value=vals[1])
    headers = ["รอบที่", "Icon", "ชื่อ", "ข้อความ", "เวลา"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="10B981")
    for i, r in enumerate(records):
        row = i + 6
        vals = [r["round"], r["icon"], r["agent"], r["message"], r["timestamp"]]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            if col == 4: cell.alignment = Alignment(wrap_text=True)
    wb.save(path)

def run_debate_stream(topic, agents_json, total_rounds, max_tokens, opening_msg, save_dir, existing_history_json):
    global debate_running_flag
    agents = json.loads(agents_json)
    if not agents or not topic.strip(): yield [], "⚠️ ข้อมูลไม่ครบ", None, None, "[]"; return
    save_dir = save_dir.strip() or "/app/downloads"
    os.makedirs(save_dir, exist_ok=True)

    records = json.loads(existing_history_json) if existing_history_json else []
    chat_history = []

    for r in records:
        text_bubble = f"**{r['icon']} {r['agent']}** ({r['timestamp']}):\n{r['message']}"
        chat_history.append([None, text_bubble])

    start_round = max([r["round"] for r in records]) + 1 if records else 1

    if not records:
        if not debate_running_flag:
            yield chat_history, "🛑 ถูกสั่งหยุด", None, None, json.dumps(records, ensure_ascii=False)
            return
        first = agents[0]
        opener = opening_msg.strip() or f"เปิดประเด็นดีเบตหัวข้อ: {topic}"
        ts = datetime.now().strftime("%H:%M:%S")
        records.append({"round": 0, "icon": first.get("icon", "🤖"), "agent": first["name"], "message": opener, "timestamp": ts})
        chat_history.append([None, f"**{first.get('icon','🤖')} {first['name']}** ({ts}):\n{opener}"])
        yield chat_history, "🟢 เริ่มต้น", None, None, json.dumps(records, ensure_ascii=False)

    consensus = False
    for rnd in range(start_round, int(total_rounds) + 1):
        if consensus: break
        for ag in agents:
            if not debate_running_flag:
                ts_label = datetime.now().strftime("%Y%m%d_%H%M%S")
                txt_path = os.path.join(save_dir, f"debate_{ts_label}.txt")
                xlsx_path = os.path.join(save_dir, f"debate_{ts_label}.xlsx")
                with open(txt_path, "w", encoding="utf-8") as f:
                    for r in records: f.write(f"[{r['timestamp']}] {r['agent']}: {r['message']}\n")
                _write_excel(records, topic, total_rounds, xlsx_path)
                yield chat_history, "🛑 หยุดสนทนาเรียบร้อย (เซฟไฟล์ชั่วคราวแล้ว)", txt_path, xlsx_path, json.dumps(records, ensure_ascii=False)
                return

            ts_start = datetime.now().strftime("%H:%M:%S")
            context = "\n".join([f"{r['agent']}: {r['message']}" for r in records[-15:]])
            current_idx = len(chat_history)

            chat_history.append([None, f"**{ag.get('icon','🤖')} {ag['name']}** ({ts_start}):\n⏳ กำลังคิด..."])
            yield chat_history, f"🔄 รอบที่ {rnd} : {ag['name']}", None, None, json.dumps(records, ensure_ascii=False)

            try:
                base = ag.get("custom_url","") if ag["provider"]=="custom" else PROVIDERS.get(ag["provider"], {"base_url": ""})["base_url"]
                client = OpenAI(api_key=ag["api_key"] or "dummy", base_url=base)

                ag_temp = float(ag.get("temperature", 0.7))
                ag_topp = float(ag.get("top_p", 0.9))

                resp = client.chat.completions.create(
                    model=ag["model"],
                    messages=[{"role": "system", "content": f"คุณคือ {ag['name']} : {ag['role']}"},
                              {"role": "user", "content": f"หัวข้อ: {topic}\nประวัติ:\n{context}\n\nตอบโต้:"}],
                    temperature=ag_temp, top_p=ag_topp, max_tokens=int(max_tokens), stream=True
                )
                txt = ""
                for chunk in resp:
                    if not debate_running_flag:
                        break
                    txt += chunk.choices[0].delta.content or ""
                    chat_history[current_idx] = [None, f"**{ag.get('icon','🤖')} {ag['name']}** ({ts_start}):\n{txt}"]
                    yield chat_history, f"⚡ {ag['name']} กำลังพูด...", None, None, json.dumps(records, ensure_ascii=False)
                reply = txt.strip()
            except Exception as e:
                reply = f"[Error: {e}]"
                chat_history[current_idx] = [None, f"**{ag.get('icon','🤖')} {ag['name']}** ({ts_start}):\n❌ {reply}"]
                yield chat_history, "⚠️ เกิดข้อผิดพลาด", None, None, json.dumps(records, ensure_ascii=False)

            if not debate_running_flag:
                break

            records.append({"round": rnd, "icon": ag.get("icon", "🤖"), "agent": ag["name"], "message": reply, "timestamp": datetime.now().strftime("%H:%M:%S")})
            if "[CONSENSUS]" in reply: consensus = True
            if consensus: break
            time.sleep(0.5)

    ts_label = datetime.now().strftime("%Y%m%d_%H%M%S")
    txt_path = os.path.join(save_dir, f"debate_{ts_label}.txt")
    xlsx_path = os.path.join(save_dir, f"debate_{ts_label}.xlsx")
    with open(txt_path, "w", encoding="utf-8") as f:
        for r in records: f.write(f"[{r['timestamp']}] {r['agent']}: {r['message']}\n")
    _write_excel(records, topic, total_rounds, xlsx_path)
    yield chat_history, "✨ เสร็จสิ้นทั้งหมด", txt_path, xlsx_path, json.dumps(records, ensure_ascii=False)

def get_table_data(agents):
    return [[i, a.get("icon", "🤖"), a.get("name", "ไม่มีชื่อ"), a.get("provider", ""), a.get("model", ""), f"Temp: {a.get('temperature', 0.7)} / TopP: {a.get('top_p', 0.9)}", a.get("role", "")[:40]+"..."] for i, a in enumerate(agents)]

def start_debate_logic():
    global debate_running_flag
    debate_running_flag = True
    return "[]"

def stop_debate_logic():
    global debate_running_flag
    debate_running_flag = False
    return "🛑 กำลังหยุดการสนทนาและสรุปผล..."

def continue_debate_logic():
    global debate_running_flag
    debate_running_flag = True
    return "⏭️ กำลังประมวลผลต่อ..."

def clear_debate_logic():
    global debate_running_flag
    debate_running_flag = False
    return [], "🗑️ ล้างจอแล้ว", "[]", None, None

with gr.Blocks(title="Multi Agent Debate v1.0", css=CUSTOM_CSS) as demo:
    agents_state = gr.State(json.dumps(DEFAULT_AGENTS, ensure_ascii=False))
    raw_history_state = gr.State("[]")
    editing_index_state = gr.State(-1)

    gr.Markdown("# 🤖 Multi Agent Debate v1.0 | <span style='font-size:18px; font-weight:normal;'>ระบบดีเบตอัจฉริยะ</span>")

    with gr.Tabs():
        with gr.Tab("💬 เวทีดีเบตหลัก (Live Arena)"):
            with gr.Row():
                with gr.Column(scale=2, elem_classes="control-box"):
                    gr.Markdown("### ⚙️ ตัวแปรควบคุมเวที")
                    topic_in = gr.Textbox(label="🎯 หัวข้อดีเบตหลัก", value="การแก้ไขน้ำท่วมใน กทม.", lines=2)
                    opening_in = gr.Textbox(label="💬 ประโยคเปิดโต๊ะ (Optional)", placeholder="ทิ้งว่างไว้ให้ระบบเปิดอัตโนมัติ")
                    with gr.Row():
                        rounds_in = gr.Number(label="🔄 รอบสูงสุด", value=10, precision=0)
                        tokens_in = gr.Number(label="📝 Max Tokens ต่อครั้ง", value=1000, precision=0)
                    with gr.Row():
                        start_btn = gr.Button("🚀 เริ่มต้นเวทีดีเบต", variant="primary")
                        stop_btn = gr.Button("🛑 หยุดสนทนาทันที", variant="stop")
                        continue_btn = gr.Button("⏭️ คุยต่อรอบถัดไป", variant="secondary")
                        clear_btn = gr.Button("🗑️ ล้างหน้าจอแชต", variant="secondary")
                with gr.Column(scale=1):
                    with gr.Row(elem_classes="stat-card"):
                        status_out = gr.Label(value="💤 Standby", label="สถานะปัจจุบัน")

            gr.Markdown("### 📢 การตอบโต้แบบเรียลไทม์")
            chatbot_ui = gr.Chatbot(label="กล่องสนทนาดีเบต", height=650)
            with gr.Row():
                dl_txt = gr.File(label="📄 Download TXT Log")
                dl_xlsx = gr.File(label="📊 Download Excel Summary")

        with gr.Tab("👥 การจัดการ Agent (Agent Settings)"):
            with gr.Row():
                with gr.Column(scale=1, elem_classes="control-box"):
                    panel_title = gr.Markdown("#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่")
                    ag_icon = gr.Dropdown(choices=EMOJI_CHOICES, value="🤖", label="Emoji Icon", allow_custom_value=True)
                    ag_name = gr.Textbox(label="ชื่อผู้พูด", placeholder="ดร.สมศักดิ์")
                    ag_provider = gr.Dropdown(choices=list(PROVIDERS.keys()), value="alibaba", label="LLM Provider")
                    ag_model = gr.Textbox(label="Model Identifier", value="qwen-plus")
                    ag_api_key = gr.Textbox(label="API Key / Token", type="password")
                    ag_base_url = gr.Textbox(label="Custom Base URL (ถ้าเลือก custom)")

                    with gr.Row():
                        ag_temp = gr.Slider(minimum=0.0, maximum=1.5, value=0.7, step=0.05, label="Temperature")
                        ag_topp = gr.Slider(minimum=0.0, maximum=1.0, value=0.9, step=0.05, label="Top P")

                    ag_role = gr.Textbox(label="Backstory / บุคลิกภาพ", lines=4)

                    with gr.Row():
                        add_agent_btn = gr.Button("💾 บันทึกข้อมูลตัวละคร", variant="primary")
                        cancel_edit_btn = gr.Button("❌ ยกเลิกการแก้ไข", variant="secondary")

                with gr.Column(scale=2):
                    gr.Markdown("#### 📋 รายชื่อผู้ร่วมสนทนาปัจจุบัน")
                    gr.Markdown("<small>💡 *เคล็ดลับ: คลิกเลือกที่แถวในตารางเพื่อดึงกลับมาแก้ไข หรือกดลบได้ทันที*</small>")
                    full_agent_table = gr.Dataframe(
                        headers=["#", "Icon", "ชื่อ", "Provider", "Model", "การตั้งค่า Parameter", "บทบาทเด่น"],
                        value=get_table_data(DEFAULT_AGENTS),
                        interactive=False
                    )
                    with gr.Row():
                        del_agent_btn = gr.Button("🗑️ ลบตัวละครที่เลือกจากตาราง", variant="stop")
                        del_all_btn = gr.Button("💥 ลบตัวละครทั้งหมด", variant="secondary")

    def on_select_cell(state, evt: gr.SelectData):
        agents = json.loads(state)
        idx = evt.index[0]
        if idx < len(agents):
            target = agents[idx]
            return (
                target.get("icon", "🤖"), target.get("name", ""), target.get("provider", "alibaba"),
                target.get("model", "qwen-plus"), target.get("api_key", ""), target.get("custom_url", ""),
                float(target.get("temperature", 0.7)), float(target.get("top_p", 0.9)), target.get("role", ""),
                f"#### 📝 กำลังแก้ไขข้อมูลลำดับที่ [{idx}]: {target.get('name')}", idx
            )
        return gr.update(), gr.update(), gr.update(), gr.update(), gr.update(), gr.update(), gr.update(), gr.update(), gr.update(), "#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่", -1

    full_agent_table.select(
        on_select_cell, inputs=[agents_state],
        outputs=[ag_icon, ag_name, ag_provider, ag_model, ag_api_key, ag_base_url, ag_temp, ag_topp, ag_role, panel_title, editing_index_state]
    )

    cancel_edit_btn.click(
        lambda: ("🤖", "", "alibaba", "qwen-plus", "", "", 0.7, 0.9, "", "#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่", -1), None,
        [ag_icon, ag_name, ag_provider, ag_model, ag_api_key, gr.State(""), ag_temp, ag_topp, ag_role, panel_title, editing_index_state]
    )

    def add_agent_logic(state, edit_idx, icon, name, provider, model, key, url, temp, topp, role):
        agents = json.loads(state)
        if not name.strip():
            return state, get_table_data(agents), icon, name, provider, model, key, url, temp, topp, role, "#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่", -1
        agent_data = {
            "icon": icon, "name": name.strip(), "provider": provider, "model": model,
            "api_key": key, "custom_url": url, "temperature": float(temp), "top_p": float(topp), "role": role
        }
        edit_idx = int(edit_idx)
        if edit_idx != -1 and edit_idx < len(agents):
            agents[edit_idx] = agent_data
        else:
            agents.append(agent_data)
        new_state = json.dumps(agents, ensure_ascii=False)
        return new_state, get_table_data(agents), "🤖", "", "alibaba", "qwen-plus", "", "", 0.7, 0.9, "", "#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่", -1

    add_agent_btn.click(
        add_agent_logic,
        [agents_state, editing_index_state, ag_icon, ag_name, ag_provider, ag_model, ag_api_key, ag_base_url, ag_temp, ag_topp, ag_role],
        [agents_state, full_agent_table, ag_icon, ag_name, ag_provider, ag_model, ag_api_key, ag_base_url, ag_temp, ag_topp, ag_role, panel_title, editing_index_state]
    )

    def del_agent_logic(state, edit_idx):
        agents = json.loads(state)
        edit_idx = int(edit_idx)
        if edit_idx == -1 or edit_idx >= len(agents):
            return state, get_table_data(agents), "🤖", "", "alibaba", "qwen-plus", "", "", 0.7, 0.9, "", "#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่", -1
        agents.pop(edit_idx)
        new_state = json.dumps(agents, ensure_ascii=False)
        return new_state, get_table_data(agents), "🤖", "", "alibaba", "qwen-plus", "", "", 0.7, 0.9, "", "#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่", -1

    del_agent_btn.click(
        del_agent_logic, inputs=[agents_state, editing_index_state],
        outputs=[agents_state, full_agent_table, ag_icon, ag_name, ag_provider, ag_model, ag_api_key, ag_base_url, ag_temp, ag_topp, ag_role, panel_title, editing_index_state]
    )

    def del_all_logic():
        empty = []
        return json.dumps(empty), get_table_data(empty), "🤖", "", "alibaba", "qwen-plus", "", "", 0.7, 0.9, "", "#### ➕ เพิ่มผู้ร่วมสนทนาคนใหม่", -1

    del_all_btn.click(
        del_all_logic, None,
        [agents_state, full_agent_table, ag_icon, ag_name, ag_provider, ag_model, ag_api_key, ag_base_url, ag_temp, ag_topp, ag_role, panel_title, editing_index_state]
    )
    ag_provider.change(lambda p: (PROVIDERS.get(p, {}).get("models") or [""])[0], ag_provider, ag_model)

    start_btn.click(start_debate_logic, None, [raw_history_state]).then(
        run_debate_stream, 
        [topic_in, agents_state, rounds_in, tokens_in, opening_in, gr.State("/app/downloads"), raw_history_state], 
        [chatbot_ui, status_out, dl_txt, dl_xlsx, raw_history_state]
    )
    
    stop_btn.click(stop_debate_logic, None, [status_out])
    
    continue_btn.click(continue_debate_logic, None, [status_out]).then(
        run_debate_stream, 
        [topic_in, agents_state, rounds_in, tokens_in, opening_in, gr.State("/app/downloads"), raw_history_state], 
        [chatbot_ui, status_out, dl_txt, dl_xlsx, raw_history_state]
    )
    
    clear_btn.click(clear_debate_logic, None, [chatbot_ui, status_out, raw_history_state, dl_txt, dl_xlsx])

# สั่งปลดล็อกอินเตอร์เฟสแชร์พอร์ตสำหรับรันบนคอนเทนเนอร์
demo.queue().launch(
    server_name="0.0.0.0", 
    server_port=7860,
    share=False,
    show_error=True
)